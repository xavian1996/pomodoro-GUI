from pomodoro_ui import Ui_MainWindow
from PyQt5 import QtCore, QtGui, QtWidgets
from openpyxl import load_workbook, Workbook
from datetime import timedelta
import re


style_sheet = """
    QTabBar::tab{
        background: rgb(43, 4, 66);
        color: black;
        border: 4px solid rgb(72, 5, 122);
        border-right-width: 0px;	
        border-radius: 10px 0;
        margin-top: 10px;
        margin-bottom: 10px;
        padding: 5px;
        padding-right: -8px
    }

    QTabBar::tab:selected  {
        background: rgb(72, 5, 122);
        color: gray;
        border-color: rgb(43, 4, 66);
    }

    QTabWidget>QWidget>QWidget{
        background: rgb(72, 5, 122);
        border: 1px solid white; border-left:0px
    }

    QCheckBox::indicator{
        width :30px;
        height :30px;
        border: 5px solid black;
        background: transparent
    }

    QCheckBox::indicator:unchecked:pressed{
        background-color : lightgreen;
    }

    QCheckBox::indicator:checked:pressed{
        background-color : #fd4e4e;
    }

    QCheckBox::indicator:checked{
        background-color : green;
    }

    
    QCheckBox::indicator:unchecked{
        background-color : red;
    }

    QRoundProgressBar {
        background-color: rgb(72, 5, 122)
        }

    QPushButton{
        background: transparent;
        border: 5px solid black;
        border-radius: 10px;
        color:white
    }
    
    QComboBox#comboBox{
        background: rgb(61, 4, 94);
        border: 5px solid black;
        border-radius: 5px;
        font-size: 20px;
        }

    QComboBox#status {
        border: 1px solid black; font-size: 10px
    }
"""
class POMODORO(QtWidgets.QWidget):
    def __init__(self, MainWindow):
        super(POMODORO, self).__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(MainWindow)
        self.inintalizeUI()

    def inintalizeUI(self):
        self.update()
        self.ui.tableWidget.itemChanged.connect(self.edit)
        self.ui.pushButton.clicked.connect(self.counterOption)
        self.ui.comboBox.setCurrentText('Select a task or, Enter new one.')
        self.working_on = None
        self.ui.comboBox.currentTextChanged.connect(self.taskOption)

        self.curent_time = 25 * 60
        self.ui.progressBar.setValue(0)
        self.ui.progressBar.setRange(0, 25 * 60)
        self.counter = 25 * 60
        self.timer = QtCore.QTimer()
        self.loop = 0

        self.timer.timeout.connect(self.handleTimer)

    def update(self):
        index = 0
        self.ui.comboBox.clear()
        for row in range(1, sheet.max_row + 1):
            if sheet['B' + str(row)].value == 'not yet':
                self.ui.comboBox.addItem(sheet['A' + str(row)].value)
            
            if row > self.ui.tableWidget.rowCount():
                self.ui.tableWidget.insertRow(self.ui.tableWidget.rowCount())
    
            for col in range(1, 5):
                cell = sheet.cell(row + 1, col).value
                if cell == None:
                    break
                if col == 2:
                    if cell:
                        status_box = QtWidgets.QComboBox()
                        if cell == 'Done':
                            status_box.addItems(['Done', 'not yet'])
                        else:
                            status_box.addItems(['not yet', 'Done'])
                        
                        status_box.setObjectName('status')
                        status_box.setProperty('row', index)
                        status_box.currentIndexChanged.connect(self.edit)
                        self.ui.tableWidget.setCellWidget(index, 1, status_box)
                
                else:
                    self.ui.tableWidget.setItem(index, col - 1, QtWidgets.QTableWidgetItem(str(cell)))

            index += 1

    def edit(self, item):
        try:
            row = item.row() + 2
            col = item.column() + 1
            data = item.text()

        except AttributeError:
            combo = self.sender()
            row = combo.property('row') + 2
            col = 2
            data = combo.currentText()
        
        sheet.cell(row, col).value = data
        #Update comboBox
        self.ui.comboBox.clear()
        for row in range(1, sheet.max_row + 1):
            if sheet['B' + str(row)].value == 'not yet':
                self.ui.comboBox.addItem(sheet['A' + str(row)].value)
        wb.save('data.xlsx')
    
    def counterOption(self):
        button = self.sender()
        if button.text() == 'START':
            self.ui.comboBox.setDisabled(True)
            self.timer.start(self.counter)
            button.setText('STOP')
        
        else:
            self.loop += 1
            self.ui.comboBox.setEnabled(True)
            self.timer.stop()
            button.setText('START')
            self.save()
            
        
    def taskOption(self):
        task = self.sender()
        self.working_on = task.currentText()

    def handleTimer(self):
        self.counter -= 1
        min = self.counter//60
        sec = int(self.counter - min * 60)
        self.ui.progressBar.setFormat(f'{min}:{sec}')
        value = self.ui.progressBar.m_value
        if value < self.curent_time:
            value = value + 1
            self.ui.progressBar.setValue(value)
        else:
            self.loop += 1
            if self.loop%4 == 0:
                self.counter = 15 * 60
                self.curent_time = self.counter
                self.ui.progressBar.setRange(0, self.curent_time)
                
            elif self.loop%2 == 1:
                self.counter = 5 * 60
                self.curent_time = self.counter
                self.ui.progressBar.setRange(0, self.curent_time)
                
            else:
                self.counter = 25 * 60
                self.ui.progressBar.setRange(0, self.curent_time)

            self.timer.start(self.counter)
            self.ui.progressBar.setValue(0)
                
    
    def save(self):
        status = lambda: 'Done' if self.ui.checkBox.isChecked() else 'not yet'
        expiry = QtCore.QDateTime.currentDateTime().toString(
            'yyyy/mm/dd hh:mm:ss'
            )
        time = str(timedelta(seconds=(25 * 60 - self.counter)))

        if sheet.max_row == 1:
            sheet['A1'] = self.working_on
            sheet['B1'] = status()
            sheet['C1'] = expiry
            sheet['D1'] = time
        
        else:
            min = (25 * 60 - self.counter) // 60; sec = int((25 * 60 - self.counter) - min * 60)
            for row in range(2, sheet.max_row + 1):
                if sheet['A' + str(row)].value == self.working_on:
                    sheet['B' + str(row)] = status()
                    sheet['C' + str(row)] = expiry
                    
                    duration = sheet['D' + str(row)].value
                    
                    try:
                        duration = re.compile(r'(\d*)\s(day|days),\s(\d*:\d*:\d*)').search(str(sheet['D' + str(row)].value))
                        days = int(duration.group(1))
                        duration = duration.group(3)
                    
                    except AttributeError:
                        days = 0
                        duration = str(sheet['D' + str(row)].value)
                    
                    lastTime = [int(i) for i in duration.split(':')]
                    lastTime = (lastTime[0] + days * 24) * 3600 +\
                        (lastTime[1] + min) * 60 + lastTime[2] + sec
                    
                    sheet['D' + str(row)] = str(timedelta(seconds=lastTime))

                    break
                elif row == sheet.max_row:
                    sheet['A' + str(row)] = self.working_on
                    sheet['B' + str(row)] = status()
                    sheet['C' + str(row)] = expiry
                    sheet['D' + str(row)] = time
        
        self.counter = 25 * 60
        return self.update()



if __name__ == "__main__":
    import sys
    
    while True:
        try: # Try if there is data.xlsx file
            wb = load_workbook('data.xlsx')
            sheet = wb.active # Select sheet
            break # Stop the loop if data.xlsx is exist

        except: # Except if data.xlsx not exist
            wb = Workbook('data.xlsx') # Create data.xlsx file
            wb.create_sheet('data') # Create sheet
            wb.save('data.xlsx') # Save the data.xlsx

    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    MainWindow.setStyleSheet(style_sheet)
    window = POMODORO(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())